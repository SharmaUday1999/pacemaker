import tkinter as tk                # python 3
from tkinter import font  as tkfont # python 3
import openpyxl
import os
import uuid 
import serial

# opening the existing excel file
filename = 'userCreds.xlsx'

if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
else:
    wb = openpyxl.Workbook()
    wb.save(filename)

# create the sheet object
ws = wb.active
wsUser = None
userFile = None

# PACEMAKER PARAMETERS
KEY = 200
LOWER_RATE_LIMIT = 0
ATRIAL_PULSE_WIDTH = 4
VENTRICULAR_PULSE_WIDTH = 6
VRP = 7
ARP = 8
ATRIAL_AMPLITUDE = 3
VENTRICULAR_AMPLITUDE = 5
AV_DELAY = 2
MESSAGE_LENGTH = 13

# PACEMAKER SERIAL PARAMETERS
BAUDRATE = 115200
TIMEOUT = 5


# TODO: We should iterate over an array and fill each column instead in the future (when/if we have time)
if ((ws['A1'] == 'Username') and (ws['B1'] == 'Password')):
    pass
else:
    ws['A1'] = 'Username'
    ws['B1'] = 'Password'
    ws['C1'] = 'User Id'
wb.save(filename)



class pacemaker(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold")
        self.geometry("1000x300")


        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)


        self.frames = {}
        for F in (welcomePage, registerPage, loginPage, mainPage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("welcomePage")

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()


class welcomePage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.winfo_toplevel().title("Pacemaker control application")

        label = tk.Label(self, text="WELCOME TO THE PACEMAKER CONTROL APPLICATION", font=controller.title_font)
        label.pack(side="top", fill="x", pady=10, padx = 20)

        button1 = tk.Button(self, text="Register",
                            command=lambda: controller.show_frame("registerPage"))
        button2 = tk.Button(self, text="Login",
                            command=lambda: controller.show_frame("loginPage"))
        button1.pack(pady=2)
        button2.pack(pady=10)


class registerPage(tk.Frame):

    def saveData(self):
        global regLabel

        if (firstnameLabelEntry.get() == "" or
            lastnameLabelEntry.get() == "" or
            emailLabelEntry.get() == "" or
            usernameLabelEntry.get() == "" or
            passwordLabelEntry.get() == "" ):

            regLabel = tk.Label(self, text = "")
            emptyLabel = tk.Label(self, text = "All inputs must be filled")
            emptyLabel.grid(row = 2, column = 4)
        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = ws.max_row
            current_column = ws.max_column
            global duplicateUsernameToggle
            duplicateUsernameToggle = 0

            for i in range(1,current_row+1):
                if ws.cell(row = i, column = 4).value == usernameLabelEntry.get():
                    duplicateUsernameToggle = 1
                    break

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location

            if current_row == 11:
                regLabel = tk.Label(self, text = "")
                regLabel = tk.Label(self, text = "Max user limit reached!")
                regLabel.grid(row = 2, column = 4)

                # clear the content of text entry box
                firstnameLabelEntry.delete(0, 'end')
                lastnameLabelEntry.delete(0, 'end')
                emailLabelEntry.delete(0, 'end')
                usernameLabelEntry.delete(0, 'end')
                passwordLabelEntry.delete(0, 'end')

            elif duplicateUsernameToggle == 0 :
                ws.cell(row=current_row + 1, column=1).value = usernameLabelEntry.get()
                ws.cell(row=current_row + 1, column=2).value = passwordLabelEntry.get()
                uniqueId = uuid.uuid4().hex[:8]
                ws.cell(row=current_row + 1, column=3).value = uniqueId

                # save the file
                wb.save(filename)

                # clear the content of text entry box
                firstnameLabelEntry.delete(0, 'end')
                lastnameLabelEntry.delete(0, 'end')
                emailLabelEntry.delete(0, 'end')
                usernameLabelEntry.delete(0, 'end')
                passwordLabelEntry.delete(0, 'end')

                regLabel = tk.Label(self, text = "")
                regLabel = tk.Label(self, text = "You have been registered!")
                regLabel.grid(row = 2, column = 4)

            elif duplicateUsernameToggle == 1:
                regLabel = tk.Label(self, text = "")
                regLabel = tk.Label(self, text = "Username not available!")
                regLabel.grid(row = 2, column = 4)

                # clear the content of text entry box
                firstnameLabelEntry.delete(0, 'end')
                lastnameLabelEntry.delete(0, 'end')
                emailLabelEntry.delete(0, 'end')
                usernameLabelEntry.delete(0, 'end')
                passwordLabelEntry.delete(0, 'end')
                duplicateUsernameToggle == 0




    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

        firstnameLabel = tk.Label(self ,text = "First Name").grid(row = 0,column = 0, padx = 5, pady = 5)
        lastnameLabel = tk.Label(self ,text = "Last Name").grid(row = 1,column = 0, padx = 5, pady = 5)
        emailLabel = tk.Label(self ,text = "Email").grid(row = 2,column = 0, padx = 5, pady = 5)
        usernameLabel = tk.Label(self ,text = "Username").grid(row = 3,column = 0, padx = 5, pady = 5)
        passwordLabel = tk.Label(self ,text = "Password").grid(row = 4,column = 0, padx = 5, pady = 5)

        global firstnameLabelEntry
        global lastnameLabelEntry
        global emailLabelEntry
        global usernameLabelEntry
        global passwordLabelEntry

        firstnameLabelEntry = tk.Entry(self)
        firstnameLabelEntry.grid(row = 0,column = 1, padx = 5, pady = 5)
        lastnameLabelEntry = tk.Entry(self)
        lastnameLabelEntry.grid(row = 1,column = 1, padx = 5, pady = 5)
        emailLabelEntry = tk.Entry(self)
        emailLabelEntry.grid(row = 2,column = 1, padx = 5, pady = 5)
        usernameLabelEntry = tk.Entry(self)
        usernameLabelEntry.grid(row = 3,column = 1, padx = 5, pady = 5)
        passwordLabelEntry = tk.Entry(self)
        passwordLabelEntry.grid(row = 4,column = 1, padx = 5, pady = 5)


        buttonRegister = tk.Button(self, text="Register", command = self.saveData)
        buttonRegister.grid(row = 6, column = 0, padx = 5, pady = 5)

        buttonReturn = tk.Button(self, text="Return to main menu",
                           command=lambda: controller.show_frame("welcomePage"))
        buttonReturn.grid(row = 6, column = 1, padx = 5, pady = 5)


class loginPage(tk.Frame):

    def login(self):
        rowMatchPassword = 1
        for i in range(1,ws.max_row+1):
            if ws.cell(row = i, column = 1).value == usernameLabelEntryLogin.get():
                rowMatchPassword = i
                break
            else:
                wrongLogin = tk.Label(self, text="Username password combination incorrect!").grid(row = 2, column = 2)


        if ws.cell(row = rowMatchPassword, column = 2).value == passwordLabelEntryLogin.get():
            
            userId = ws.cell(row = rowMatchPassword, column = 3).value
            global userFile
            userFile = userId + '.xlsx'
            if os.path.isfile(userFile):
                wb = openpyxl.load_workbook(userFile)
            else:
                wb = openpyxl.Workbook()
                wb.save(userFile)

            # create the sheet object
            global wsUser
            wsUser = wb.active

            
            # TODO: We should iterate over an array and fill each column instead in the future (when/if we have time)
            if ((wsUser['A1'] == 'Lower Rate Limit') and (wsUser['B1'] == 'Upper Rate Limit') and (wsUser['C1'] == 'AV Delay') and (wsUser['D1'] == 'Atrial Amplitude') and (wsUser['E1'] == 'Atrial Pulse Width')
                and (wsUser['F1'] == 'Ventricular Amplitude') and (wsUser['G1'] == 'Ventricular Pulse Width') and (wsUser['H1'] == 'VRP') and (wsUser['I1'] == 'ARP')):
                pass
            else:
                wsUser['A1'] = 'Lower Rate Limit'
                wsUser['B1'] = 'Upper Rate Limit'
                wsUser['C1'] = 'AV Delay'
                wsUser['D1'] = 'Atrial Amplitude'
                wsUser['E1'] = 'Atrial Pulse Width'
                wsUser['F1'] = 'Ventricular Amplitude'
                wsUser['G1'] = 'Ventricular Pulse Width'
                wsUser['H1'] = 'VRP'
                wsUser['I1'] = 'ARP'
            wb.save(userFile)    
            loggedInRow = rowMatchPassword
            self.controller.show_frame('mainPage')
            usernameLabelEntryLogin.delete(0,'end')
            passwordLabelEntryLogin.delete(0,'end')
            mainPage.setLoggedInRow(self.controller.frames['mainPage'], loggedInRow)

        #implement incorrect login text


    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="Login", font=controller.title_font).grid(row = 0, column = 0)


        usernameLabel = tk.Label(self ,text = "Username").grid(row = 1,column = 0, padx = 5, pady = 5)
        passwordLabel = tk.Label(self ,text = "Password").grid(row = 2,column = 0, padx = 5, pady = 5)

        global usernameLabelEntryLogin
        global passwordLabelEntryLogin

        usernameLabelEntryLogin = tk.Entry(self)
        usernameLabelEntryLogin.grid(row = 1,column = 1, padx = 5, pady = 5)
        passwordLabelEntryLogin = tk.Entry(self)
        passwordLabelEntryLogin.grid(row = 2,column = 1, padx = 5, pady = 5)

        buttonLogin = tk.Button(self, text="Login", command = self.login)
        buttonLogin.grid(row = 6, column = 0, padx = 5, pady = 5)

        buttonReturn = tk.Button(self, text="Return to welcome page",
                           command=lambda: controller.show_frame("welcomePage"))
        buttonReturn.grid(row = 6, column = 1, padx = 5, pady = 5)

class mainPage(tk.Frame):
    _state = 'none'
    # TODO: Should move this into a separate module, it's just an object storing constants
    _pacingModes = {
        'AOO' : 0,
        'VOO' : 1,
        'AAI' : 2,
        'VVI' : 3,
        'DOO' : 4,
    }
    _params = {
        'LOWER_RATE_LIMIT' : {
            'AOO': 'normal',
            'VOO': 'normal',
            'AAI' : 'normal',
            'VVI' : 'normal',
            'DOO' : 'normal'
            },
        'UPPER_RATE_LIMIT' : {
            'AOO': 'normal',
            'VOO': 'normal',
            'AAI' : 'normal',
            'VVI' : 'normal',
            'DOO' : 'normal'
            },
        'FIXED_AV_DELAY' : {
            'AOO': 'disabled',
            'VOO': 'disabled',
            'AAI' : 'disabled',
            'VVI' : 'disabled',
            'DOO' : 'normal'
        },
        'ATRIAL_AMPLITUDE' : {
            'AOO': 'normal',
            'VOO': 'disabled',
            'AAI' : 'normal',
            'VVI' : 'disabled',
            'DOO' : 'normal'
            },
        'ATRIAL_PULSE_WIDTH': {
            'AOO': 'normal',
            'VOO': 'disabled',
            'AAI' : 'normal',
            'VVI' : 'disabled',
            'DOO' : 'normal'
            },
        'VENTRICULAR_AMPLITUDE': {
            'AOO': 'disabled',
            'VOO': 'normal',
            'AAI' : 'disabled',
            'VVI' : 'normal',
            'DOO' : 'normal'
            },
        'VENTRICULAR_PULSE_WIDTH': {
            'AOO': 'disabled',
            'VOO': 'normal',
            'AAI' : 'disabled',
            'VVI' : 'normal',
            'DOO' : 'normal'
            },    
        'VRP' : {
            'AOO': 'disabled',
            'VOO': 'disabled',
            'AAI' : 'disabled',
            'VVI' : 'normal',
            'DOO' : 'disabled'
            },
        'ARP' : {
            'AOO': 'disabled',
            'VOO': 'disabled',
            'AAI' : 'normal',
            'VVI' : 'disabled',
            'DOO' : 'disabled'
            }
        }
    _paramBoundaries = {
        'LRL' : {
            'min' : 30.0,
            'max' : 175.0,
            'exception': 30.0,
            },
        'URL' : {
            'min' : 50.0,
            'max' : 175.0,
            'exception': 50.0,
            },    
        'avDelay' : {
            'min' : 70.0,
            'max' : 300.0,
            'exception': 70.0
            },    
        'atrialAmp' : {
            'min' : 0.5,
            'max' : 7.0,
            'exception': 0.5,
            }, 
        'atrialPW' : {
            'min' : 0.1,
            'max' : 1.9,
            'exception': 0.05,
            },    
        'venAmp' : {
            'min' : 0.5,
            'max' : 7.0,
            'exception': 0.5,
            },  
        'venPW' : {
            'min' : 0.1,
            'max' : 1.9,
            'exception': 0.05,
            }, 
        'VRP' : {
            'min' : 150.0,
            'max' : 500.0,
            'exception': 150.0,
            }, 
        'ARP' : {
            'min' : 150,
            'max' : 500,
            'exception': 150.0,
            },               
    }

    def _onTouch(self, event):
        
        lrlValue = float(lrlEntry.get())
        urlValue = float(urlEntry.get())
        avDelayValue = float(avDelayEntry.get())
        atrialAmpValue = float(atrialAmpEntry.get())
        atrialPWValue = float(atrialPWEntry.get())
        venAmpValue = float(venAmpEntry.get())
        venPWValue = float(venPWEntry.get())
        vrpValue = float(vrpEntry.get())
        arpValue = float(arpEntry.get())


        if (lrlValue >= self._paramBoundaries['LRL']['min'] and lrlValue <= self._paramBoundaries['LRL']['max'] or lrlValue == self._paramBoundaries['LRL']['exception']) :
            lrlEntry.configure({'background' : 'white'})
        else :
            lrlEntry.configure({'background' : '#ff6b6b'})

        if (urlValue >= self._paramBoundaries['URL']['min'] and urlValue <= self._paramBoundaries['URL']['max'] or urlValue == self._paramBoundaries['URL']['exception']) : #FIX THIS FOR ALL
            urlEntry.configure({'background' : 'white'})
        else :
            urlEntry.configure({'background' : '#ff6b6b'})

        if (avDelayValue >= self._paramBoundaries['avDelay']['min'] and avDelayValue <= self._paramBoundaries['avDelay']['max'] or avDelayValue == self._paramBoundaries['avDelay']['exception']) :
            avDelayEntry.configure({'background' : 'white'})
        else :
            avDelayEntry.configure({'background' : '#ff6b6b'})

        if (atrialAmpValue >= self._paramBoundaries['atrialAmp']['min'] and atrialAmpValue <= self._paramBoundaries['atrialAmp']['max'] or atrialAmpValue == self._paramBoundaries['atrialAmp']['exception']) :
            atrialAmpEntry.configure({'background' : 'white'})
        else :
            atrialAmpEntry.configure({'background' : '#ff6b6b'})

        if (atrialPWValue >= self._paramBoundaries['atrialPW']['min'] and atrialPWValue <= self._paramBoundaries['atrialPW']['max'] or atrialPWValue == self._paramBoundaries['atrialPW']['exception']) :
            atrialPWEntry.configure({'background' : 'white'})
        else :
            atrialPWEntry.configure({'background' : '#ff6b6b'})

        if (venAmpValue >= self._paramBoundaries['venAmp']['min'] and venAmpValue <= self._paramBoundaries['venAmp']['max'] or venAmpValue == self._paramBoundaries['venAmp']['exception']) :
            venAmpEntry.configure({'background' : 'white'})
        else :
            venAmpEntry.configure({'background' : '#ff6b6b'})

        if (venPWValue >= self._paramBoundaries['venPW']['min'] and venPWValue <= self._paramBoundaries['venPW']['max'] or venPWValue == self._paramBoundaries['venPW']['exception']) :
            venPWEntry.configure({'background' : 'white'})
        else :
            venPWEntry.configure({'background' : '#ff6b6b'})

        if (vrpValue >= self._paramBoundaries['VRP']['min'] and vrpValue <= self._paramBoundaries['VRP']['max'] or vrpValue == self._paramBoundaries['VRP']['exception']) :
            vrpEntry.configure({'background' : 'white'})
        else :
            vrpEntry.configure({'background' : '#ff6b6b'})

        if (arpValue >= self._paramBoundaries['ARP']['min'] and arpValue <= self._paramBoundaries['ARP']['max'] or arpValue == self._paramBoundaries['ARP']['exception']) :
            arpEntry.configure({'background' : 'white'})
        else :
            arpEntry.configure({'background' : '#ff6b6b'})            


    def setMode(self, mode):
        #Change state and fields depending on which pacing mode is selected. Default none

        self._state = mode
        self._loggedInRow = 1

        lrlEntry.configure(state= self._params['LOWER_RATE_LIMIT'][mode])
        urlEntry.configure(state= self._params['UPPER_RATE_LIMIT'][mode])
        avDelayEntry.configure(state = self._params['FIXED_AV_DELAY'][mode])
        atrialAmpEntry.configure(state= self._params['ATRIAL_AMPLITUDE'][mode])
        atrialPWEntry.configure(state= self._params['ATRIAL_PULSE_WIDTH'][mode])
        venAmpEntry.configure(state= self._params['VENTRICULAR_AMPLITUDE'][mode])
        venPWEntry.configure(state= self._params['VENTRICULAR_PULSE_WIDTH'][mode])
        vrpEntry.configure(state= self._params['VRP'][mode])
        arpEntry.configure(state= self._params['ARP'][mode])
        # This is ugly, but .set didn't seem to work
        selectedLabel = tk.Label(self ,text = mode).grid(row = 22,column = 1)

    def Save(self):

        global errorLabel
    
        if os.path.isfile(userFile):
            wb = openpyxl.load_workbook(userFile)

        # create the sheet object
        wsUser = wb.active

        if (float(lrlEntry.get()) >= float(30) and float(lrlEntry.get()) <=float(175)) and (float(urlEntry.get()) >= 50.0 and float(urlEntry.get()) <=175.0) and (float(atrialAmpEntry.get()) >= 0.5 and float(atrialAmpEntry.get()) <= 7.0) and (float(atrialPWEntry.get()) == 0.05 or float(atrialPWEntry.get()) >=0.1 and float(atrialPWEntry.get())<= 1.9) and (float(venAmpEntry.get()) >= 0.5 and float(venAmpEntry.get()) <=7.0) and (float(venPWEntry.get()) == 0.05 or float(venPWEntry.get()) >=0.1 and float(venPWEntry.get())<= 1.9) and (float(vrpEntry.get()) >= 150.0 and float(vrpEntry.get()) <=500.0) and  (float(arpEntry.get()) >= 150.0 and float(arpEntry.get()) <=500.0):
            wsUser.cell(row = 2, column = 1).value = lrlEntry.get()
            wsUser.cell(row = 2, column = 2).value = urlEntry.get()
            wsUser.cell(row = 2, column = 3).value = avDelayEntry.get()
            wsUser.cell(row = 2, column = 4).value = atrialAmpEntry.get()
            wsUser.cell(row = 2, column = 5).value = atrialPWEntry.get()
            wsUser.cell(row = 2, column = 6).value = venAmpEntry.get()
            wsUser.cell(row = 2, column = 7).value = venPWEntry.get()
            wsUser.cell(row = 2, column = 8).value = vrpEntry.get()
            wsUser.cell(row = 2, column = 9).value = arpEntry.get()
            errorLabel = tk.Label(self ,text = "Values Saved",)
            errorLabel.grid(row = 19, column = 3, padx = 5, pady = 5)
        else:
            errorLabel = tk.Label(self ,text = "Error in values",)
            errorLabel.grid(row = 19, column = 3, padx = 5, pady = 5)
        self._onTouch(None)

        wb.save(userFile)

    def populateUserData(self):

        #Clear any data that might exist
        lrlEntry.delete(0, 'end')
        urlEntry.delete(0, 'end')
        avDelayEntry.delete(0, 'end')
        atrialAmpEntry.delete(0, 'end')
        atrialPWEntry.delete(0, 'end')
        venAmpEntry.delete(0, 'end')
        venPWEntry.delete(0, 'end')
        vrpEntry.delete(0, 'end')
        arpEntry.delete(0, 'end')

        lrlEntry.insert(0, wsUser.cell(row = 2, column = 1).value if type(wsUser.cell(row = 2, column = 1).value) == str else 0)
        urlEntry.insert('end', wsUser.cell(row = 2, column = 2).value if type(wsUser.cell(row = 2, column = 2).value) == str else 0)
        avDelayEntry.insert('end', wsUser.cell(row = 2, column = 3).value if type(wsUser.cell(row = 2, column = 3).value) == str else 0)
        atrialAmpEntry.insert('end', wsUser.cell(row = 2, column = 4).value if type(wsUser.cell(row = 2, column = 4).value) == str else 0)
        atrialPWEntry.insert('end', wsUser.cell(row = 2, column = 5).value if type(wsUser.cell(row = 2, column = 5).value) == str else 0)
        venAmpEntry.insert('end', wsUser.cell(row = 2, column = 6).value if type(wsUser.cell(row = 2, column = 6).value) == str else 0)
        venPWEntry.insert('end', wsUser.cell(row = 2, column = 7).value if type(wsUser.cell(row = 2, column = 7).value) == str else 0)
        vrpEntry.insert('end', wsUser.cell(row = 2, column = 8).value if type(wsUser.cell(row = 2, column = 8).value) == str else 0)
        arpEntry.insert('end', wsUser.cell(row = 2, column = 9).value if type(wsUser.cell(row = 2, column = 9).value) == str else 0)

        self.setMode('AOO')

    def setLoggedInRow(self, row):
        self._loggedInRow = row
        self.populateUserData()

    ''' PARAMETER PROCESSING '''

    # lower rate lim
    def LRL(self, num):
        num = float(num)
        return int(num)

    # atrial pulse width
    def APW(self, num):
        num = float(num)
        return int(num*100)

    # ventricular pulse width
    def VPW(self, num):
        num = float(num)
        return int(num*100)
        
    # VRP (split into two summands)
    def fVRP(self, num):
        num = float(num)
        remainder = 1 if (num % 2 == 1) else 0
        return [int(num//2), int(num//2 + remainder)]

    # ARP
    def fARP(self, num):
        num = float(num)
        remainder = 1 if (num % 2 == 1) else 0
        return [int(num//2), int(num//2 + remainder)]

    # atrial amplitude
    def AA(self, num):
        num = float(num)
        return int(num*10)

    # ventricular amplitude
    def VA(self, num):
        num = float(num)
        return int(num*10)

    # AV delay
    def AVD(self, num):
        num = float(num)
        remainder = 1 if (num % 2 == 1) else 0
        summand1 = int(num//2)
        summand2 = int(num//2) + remainder
        return [summand1, summand2]

    
    def _connect(self):
        print('just press SEND')


    def _send(self):
        output = [0] * MESSAGE_LENGTH
        output[0] = KEY
        output[1] = self._pacingModes[self._state]
        print (wsUser.cell(row = 2, column = ATRIAL_PULSE_WIDTH + 1).value)
        print (wsUser.cell(row = 2, column = VENTRICULAR_PULSE_WIDTH + 1).value)
        print (wsUser.cell(row = 2, column = VRP + 1).value)
        print (wsUser.cell(row = 2, column = ARP + 1).value)
        print (wsUser.cell(row = 2, column = ATRIAL_AMPLITUDE + 1).value)
        print (wsUser.cell(row = 2, column = VENTRICULAR_AMPLITUDE + 1).value)
        print (wsUser.cell(row = 2, column = AV_DELAY + 1).value)
        output[2] = self.LRL(wsUser.cell(row = 2, column = LOWER_RATE_LIMIT + 1).value)
        output[3] = self.APW(wsUser.cell(row = 2, column = ATRIAL_PULSE_WIDTH + 1).value)
        output[4] = self.VPW(wsUser.cell(row = 2, column = VENTRICULAR_PULSE_WIDTH + 1).value)
        output[5] = self.fVRP(wsUser.cell(row = 2, column = VRP + 1).value)[0]
        output[6] = self.fVRP(wsUser.cell(row = 2, column = VRP + 1).value)[1]
        output[7] = self.fARP(wsUser.cell(row = 2, column = ARP + 1).value)[0]
        output[8] = self.fARP(wsUser.cell(row = 2, column = ARP + 1).value)[1]
        output[9] = self.AA(wsUser.cell(row = 2, column = ATRIAL_AMPLITUDE + 1).value) 
        output[10] = self.VA(wsUser.cell(row = 2, column = VENTRICULAR_AMPLITUDE + 1).value)
        output[11] = self.AVD(wsUser.cell(row = 2, column = AV_DELAY + 1).value)[0]
        output[12] = self.AVD(wsUser.cell(row = 2, column = AV_DELAY + 1).value)[1]
        print(output)  
        print('serial port:' + serPortEntry.get())
        try:
            
            ser = serial.Serial(
                
                port = serPortEntry.get(),
                baudrate = BAUDRATE,
                timeout = TIMEOUT
            )
            ser.write(bytes(output))
            print('data seeeent')
            successLabel = tk.Label(self ,text = "Values sent", fg = "green")
            successLabel.grid(row = 19, column = 3, padx = 5, pady = 5)
        except:   
            print("serial failure")
            successLabel = tk.Label(self ,text = "Values did not send", fg = "red")
            successLabel.grid(row = 19, column = 3, padx = 5, pady = 5)


    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        label = tk.Label(self, text="WELCOME TO THE PACEMAKER PORTAL", font=controller.title_font).grid(row = 0, column = 0, columnspan = 5)

        
        #Page Navigation
        logoutButton = tk.Button(self, text="Logout",
                           command=lambda: controller.show_frame('welcomePage'))
        logoutButton.grid(row = 20, column = 3, padx = 5, pady = 5)

        #Pacing Modes
        stateLabel = tk.Label(self ,text = "State").grid(row = 22,column = 0)
        selectedLabel = tk.Label(self ,text = "None").grid(row = 22,column = 1)

        aooButton = tk.Button(self, text="AOO",
                           command=lambda: self.setMode('AOO'))
        aooButton.grid(row = 10, column = 0, padx = 5, pady = 5)

        vooButton = tk.Button(self, text="VOO",
                           command=lambda: self.setMode('VOO'))
        vooButton.grid(row = 10, column = 2, padx = 5, pady = 5)

        aaiButton = tk.Button(self, text="AAI",
                           command=lambda: self.setMode('AAI'))
        aaiButton.grid(row = 10, column = 4, padx = 5, pady = 5)

        vviButton = tk.Button(self, text="VVI",
                           command=lambda: self.setMode('VVI'))
        vviButton.grid(row = 10, column = 6, padx = 5, pady = 5)

        dooButton = tk.Button(self, text="DOO",
                           command=lambda: self.setMode('DOO'))
        dooButton.grid(row = 10, column = 8, padx = 5, pady = 5)


        #Pacing Mode Parameters
        lrlLabel = tk.Label(self ,text = "Lower Rate Limit (ppm): 30 - 175",).grid(row = 15,column = 0, padx = 1, pady = 1, columnspan=2)
        urlLabel = tk.Label(self ,text = "Upper Rate Limit (ppm): 50 - 175").grid(row = 16,column = 0, padx = 1, pady = 1, columnspan=2)
        avDelayLabel = tk.Label(self ,text = "AV Delay (ms): 70 - 300",).grid(row = 17,column = 0, padx = 1, pady = 1, columnspan=2)
        atrialAmpLabel = tk.Label(self ,text = "Atrial Amplitude (V): 0.5 - 7.0").grid(row = 18,column = 0, padx = 1, pady = 1, columnspan=2)
        atrialPWLabel = tk.Label(self ,text = "Atrial Pulse Width (ms): 0.05 or 0.1 - 1.9").grid(row = 19,column = 0, padx = 1, pady = 1, columnspan=2)
        venAmpLabel = tk.Label(self ,text = "Ventricular Amplitude (V): 0.5 - 7.0").grid(row = 15,column = 4, padx = 1, pady = 1, columnspan=2)
        venPWLabel = tk.Label(self ,text = "Ventricular Pulse Width (ms): 0.05 or 0.1 - 1.9",).grid(row = 16,column = 4, padx = 1, pady = 1, columnspan=2)
        vrpLabel = tk.Label(self ,text = "VRP (ms): 150 - 500",).grid(row = 17,column = 4, padx = 1, pady = 1, columnspan=2)
        arpLabel = tk.Label(self ,text = "ARP (ms): 150 - 500",).grid(row = 18,column = 4, padx = 1, pady = 1, columnspan=2)


        global lrlEntry
        global urlEntry
        global avDelayEntry
        global atrialAmpEntry
        global atrialPWEntry
        global venAmpEntry
        global venPWEntry
        global vrpEntry
        global arpEntry
        global serPortEntry


        lrlEntry = tk.Entry(self, width=5, disabledbackground='grey')
        lrlEntry.grid(row = 15,column = 2)
        urlEntry = tk.Entry(self, width=5, disabledbackground='grey')
        urlEntry.grid(row = 16,column = 2, padx = 1, pady = 1)
        avDelayEntry = tk.Entry(self, width=5, disabledbackground='grey')
        avDelayEntry.grid(row = 17,column = 2, padx = 1, pady = 1)
        atrialAmpEntry = tk.Entry(self, width=5, disabledbackground='grey')
        atrialAmpEntry.grid(row = 18,column = 2, padx = 1, pady = 1)
        atrialPWEntry = tk.Entry(self, width=5, disabledbackground='grey')
        atrialPWEntry.grid(row = 19,column = 2, padx = 1, pady = 1)
        venAmpEntry = tk.Entry(self, width=5, disabledbackground='grey')
        venAmpEntry.grid(row = 15,column = 7, padx = 1, pady = 1)
        venPWEntry = tk.Entry(self, width=5, disabledbackground='grey')
        venPWEntry.grid(row = 16,column = 7, padx = 1, pady = 1)
        vrpEntry = tk.Entry(self, width=5, disabledbackground='grey')
        vrpEntry.grid(row = 17,column = 7, padx = 1, pady = 1)
        arpEntry = tk.Entry(self, width=5, disabledbackground='grey')
        arpEntry.grid(row = 18,column = 7, padx = 1, pady = 1)

        # Serial Communication
        serPortComLabel = tk.Label(self ,text = "Serial Port Communication",).grid(row = 21,column = 5)
        serPortLabel = tk.Label(self ,text = "Port").grid(row = 22,column = 4)
        serPortEntry = tk.Entry(self, width=5)
        serPortEntry.grid(row = 22, column = 5)
        serPortEntry.insert('end', '/dev/cu.usbmodem0000001234561')
        eightBitKeyLabel = tk.Label(self ,text = "8-bit key").grid(row = 22,column = 6)
        eightBitKeyEntry = tk.Entry(self, width=5)
        eightBitKeyEntry.grid(row = 22, column = 7)
        sendButton = tk.Button(self, text="Send", command= self._send)
        sendButton.grid(row = 22, column = 9)


        lrlEntry.bind('<FocusOut>', self._onTouch)
        urlEntry.bind('<FocusOut>', self._onTouch)
        avDelayEntry.bind('<FocusOut>', self._onTouch)
        atrialAmpEntry.bind('<FocusOut>', self._onTouch)
        atrialPWEntry.bind('<FocusOut>', self._onTouch)
        venAmpEntry.bind('<FocusOut>', self._onTouch)
        venPWEntry.bind('<FocusOut>', self._onTouch)
        vrpEntry.bind('<FocusOut>', self._onTouch)
        arpEntry.bind('<FocusOut>', self._onTouch)

        buttonSave = tk.Button(self, text="Save", command = self.Save)
        buttonSave.grid(row = 19, column = 6, padx = 5, pady = 5)



if __name__ == "__main__":
    app = pacemaker()
    app.mainloop()
